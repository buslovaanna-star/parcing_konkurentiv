import streamlit as st
import pandas as pd
import numpy as np
import io, re
from pathlib import Path
from datetime import datetime

st.set_page_config(page_title="ProcureAI", page_icon="🛒", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Unbounded:wght@700;800&family=Inter:wght@400;500;600&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif;}
.app-header{background:linear-gradient(135deg,#080f1a 0%,#0d1e30 100%);border:1px solid #1a3045;border-radius:16px;padding:28px 36px;margin-bottom:20px;position:relative;overflow:hidden;}
.app-header::after{content:'';position:absolute;top:-60px;right:-60px;width:220px;height:220px;border-radius:50%;background:radial-gradient(circle,#0ea5e915 0%,transparent 70%);}
.app-header h1{font-family:'Unbounded',sans-serif;font-size:1.7rem;font-weight:800;color:#eef6ff;margin:0 0 4px 0;}
.app-header p{color:#5b8aa8;font-size:.88rem;margin:0;}
.badge{display:inline-block;background:#0ea5e912;color:#0ea5e9;border:1px solid #0ea5e928;border-radius:20px;padding:2px 10px;font-size:.7rem;font-family:'DM Mono',monospace;margin-right:6px;}
.kpi-row{display:flex;gap:12px;flex-wrap:wrap;margin:16px 0;}
.kpi{flex:1;min-width:140px;background:#080f1a;border:1px solid #1a3045;border-radius:12px;padding:14px 16px;position:relative;}
.kpi::after{content:'';position:absolute;bottom:0;left:0;right:0;height:2px;border-radius:0 0 12px 12px;}
.kpi.sky::after{background:#0ea5e9;}.kpi.grn::after{background:#22c55e;}.kpi.amb::after{background:#f59e0b;}.kpi.red::after{background:#ef4444;}.kpi.vio::after{background:#a855f7;}
.kpi .l{color:#4d7a96;font-size:.68rem;font-weight:600;letter-spacing:.5px;text-transform:uppercase;margin-bottom:5px;}
.kpi .v{color:#eef6ff;font-family:'Unbounded',sans-serif;font-size:1.55rem;font-weight:800;line-height:1.1;}
.kpi .s{color:#2d5570;font-size:.7rem;font-family:'DM Mono',monospace;margin-top:2px;}
.sup-row{display:flex;gap:12px;flex-wrap:wrap;margin:12px 0;}
.sc{flex:1;min-width:180px;border-radius:12px;padding:16px 20px;border:1px solid;}
.sc.ih{background:#04101a;border-color:#0ea5e925;}.sc.vw{background:#04100a;border-color:#22c55e25;}.sc.dsn{background:#100a04;border-color:#f59e0b25;}.sc.atl{background:#0d0414;border-color:#a855f725;}
.sc .sn{font-size:.68rem;font-weight:700;letter-spacing:1px;text-transform:uppercase;font-family:'DM Mono',monospace;margin-bottom:5px;}
.sc.ih .sn{color:#0ea5e9;}.sc.vw .sn{color:#22c55e;}.sc.dsn .sn{color:#f59e0b;}.sc.atl .sn{color:#a855f7;}
.sc .sv{color:#eef6ff;font-family:'Unbounded',sans-serif;font-size:1.3rem;font-weight:800;margin:4px 0 2px;}
.sc .sd{color:#4d7a96;font-size:.75rem;}
.sh{display:flex;align-items:center;gap:8px;border-bottom:1px solid #1a3045;padding-bottom:7px;margin:20px 0 12px;}
.sh h3{color:#eef6ff;font-size:.92rem;font-weight:600;margin:0;}
.dot{width:6px;height:6px;border-radius:50%;flex-shrink:0;}
section[data-testid="stSidebar"]>div{background:#050d16 !important;}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="app-header">
  <span class="badge">PROCURE</span><span class="badge">AI v3</span>
  <h1>🛒 ProcureAI — Аналіз закупівель</h1>
  <p>Сайт + Магазини → оптимальне замовлення з аналізом маржинальності</p>
</div>""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Налаштування")
    st.divider()
    usd_rate = st.number_input("💱 Курс USD → UAH", 1.0, 200.0, 41.5, 0.5)
    disp_cur = st.radio("🪙 Валюта", ["UAH","USD"])
    cur      = "₴" if disp_cur=="UAH" else "$"
    st.divider()
    st.markdown("**📦 Замовлення**")
    horizon   = st.slider("Горизонт закупівлі (днів)", 7, 90, 30, 7)
    lead_time = st.slider("Lead time (днів)", 0, 60, 14, 1)
    safety    = st.slider("Страховий запас (днів)", 0, 60, 30, 5)
    total_days = horizon + lead_time + safety
    st.caption(f"Покриття: **{total_days} днів**")
    st.divider()
    st.markdown("**📊 Фільтри продажів**")
    sales_mode    = st.radio("Базовий період", ["Весь період","Останні N місяців"])
    recent_months = st.number_input("N місяців", 1, 17, 3, 1) if sales_mode=="Останні N місяців" else None
    min_rent      = st.slider("Ігнорувати продажі з рентабельністю нижче %", 0, 60, 30, 5,
                              help="Акційні продажі не враховуються в потребі")
    st.divider()
    st.markdown("**💰 Маржинальність**")
    margin_min = st.slider("Мінімальна маржа закупівлі %", 0, 80, 25, 5)
    only_avail = st.checkbox("Тільки товари в наявності", True)
    st.divider()
    st.markdown("**🏆 ABC**")
    abc_inc = st.multiselect("Включити категорії", ["A","B","C"], default=["A","B","C"])

# ── Helpers ───────────────────────────────────────────────────────
def cs(s):
    return '' if pd.isna(s) else str(s).replace('\xa0',' ').strip()

def to_num(v):
    if pd.isna(v): return 0.0
    s = str(v).replace('\xa0','').replace(' ','').replace(',','.')
    try: return float(s)
    except: return 0.0

def read_f(f):
    if f is None: return None
    if f.name.endswith('.csv'): return pd.read_csv(f)
    eng = 'xlrd' if f.name.endswith('.xls') else 'openpyxl'
    return pd.read_excel(f, engine=eng)

@st.cache_data(show_spinner=False)
def cached_read_f(file_bytes, file_name):
    """Кешована версія read_f — уникає повторного парсингу Excel/CSV при ререндері."""
    if file_name.endswith('.csv'):
        return pd.read_csv(io.BytesIO(file_bytes))
    eng = 'xlrd' if file_name.endswith('.xls') else 'openpyxl'
    return pd.read_excel(io.BytesIO(file_bytes), engine=eng)

@st.cache_data(show_spinner=False)
def cached_read_excel_sheet(file_bytes, sheet_name, header, engine=None):
    """Кешує читання конкретного листа Excel-файлу за хешем його байтів.
    При повторному рендері Streamlit (зміна слайдера тощо) той самий файл
    не перечитується з диска — береться готовий DataFrame з кешу."""
    return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=header, engine=engine)

@st.cache_data(show_spinner=False)
def cached_excel_sheet_names(file_bytes):
    return pd.ExcelFile(io.BytesIO(file_bytes)).sheet_names

def find_col(df, kws):
    for c in df.columns:
        if any(k in str(c).lower() for k in kws): return c
    return None

def is_avail_ih(x):
    return False if pd.isna(x) else ('наявн' in str(x).lower() or 'в налич' in str(x).lower())

def is_avail_dsn(x):
    return False if pd.isna(x) else ('в наявності' == str(x).strip().lower() or 'в налич' in str(x).lower())

SKIP_ARTS  = {f'{i}*' for i in range(1,15)}
SKIP_NAMES = {'доставка','нова пошта','новая почта','iherb','магазини'}

def parse_sales(df_raw, step, data_row, min_rent_pct, n_months=None):
    qty_off  = 4 if step==7 else 2
    rent_off = 5 if step==7 else 3
    arr = df_raw.values  # numpy масив — набагато швидше за .iloc у циклі
    months = []
    for c in range(0, arr.shape[1], step):
        val = cs(arr[1, c])
        m = re.sub(r'Період[:\s]*', '', val).replace('р.','').strip()
        if m: months.append((c, m))
    if n_months: months = months[-n_months:]
    data_block = arr[data_row:, :]  # відрізаємо службові рядки одним зрізом
    records = {}
    for col_s, mname in months:
        names = data_block[:, col_s]
        arts  = data_block[:, col_s+1]
        qtys  = data_block[:, col_s+qty_off]
        rents = data_block[:, col_s+rent_off]
        for name_raw, art_raw, qty_raw, rent_raw in zip(names, arts, qtys, rents):
            art = cs(art_raw)
            if not art or art in SKIP_ARTS: continue
            name = cs(name_raw)
            if name.lower() in SKIP_NAMES: continue
            qty = to_num(qty_raw)
            if qty <= 0: continue
            rent = to_num(rent_raw)
            if rent > 0 and rent < min_rent_pct: continue
            if art not in records: records[art] = {'Назва': name, 'міс': {}}
            records[art]['міс'][mname] = records[art]['міс'].get(mname, 0) + qty
    if not records:
        return pd.DataFrame(columns=['Артикул_IH','Назва','Продано_всього','Місяців','Середньодень'])
    rows = []
    n_m = len(months)
    for art, v in records.items():
        total = sum(v['міс'].values())
        rows.append({'Артикул_IH':art, 'Назва':v['Назва'],
                     'Продано_всього':total, 'Місяців':n_m,
                     'Середньодень': total/(n_m*30) if n_m>0 else 0})
    return pd.DataFrame(rows)

def parse_availability_days(df_raw, n_months=None):
    """
    Парсить вкладку 'Залишки' сайту: крок 4 колонки (Артикул/Номенклатура/Днів/NaN),
    дані з рядка 4. Повертає суму фактичних днів наявності по кожному артикулу
    за обраний період (весь / останні N місяців) — для точного розрахунку
    середньоденних продажів замість ділення на календарні 30 днів/місяць.
    """
    step = 4
    arr = df_raw.values
    months = []
    for c in range(0, arr.shape[1], step):
        val = cs(arr[1, c])
        m = re.sub(r'Період[:\s]*', '', val).replace('р.','').strip()
        if m: months.append(c)
    if n_months: months = months[-n_months:]
    data_block = arr[4:, :]
    days_sum = {}
    for col_s in months:
        arts = data_block[:, col_s]
        days_col = data_block[:, col_s+2]
        for art_raw, days_raw in zip(arts, days_col):
            art = cs(art_raw)
            if not art or art in SKIP_ARTS: continue
            days_sum[art] = days_sum.get(art, 0) + to_num(days_raw)
    if not days_sum:
        return pd.DataFrame(columns=['Артикул_IH','Днів_наявності'])
    return pd.DataFrame([{'Артикул_IH':a, 'Днів_наявності':d} for a,d in days_sum.items()])

def parse_stock(df_raw, data_row):
    arr = df_raw.values
    data_block = arr[data_row:, :]
    has_transit_col = arr.shape[1] > 3
    rows = []
    for r in data_block:
        art = cs(r[1])
        if not art or art in SKIP_ARTS: continue
        q = max(to_num(r[2]), 0)
        in_transit = max(to_num(r[3]), 0) if has_transit_col else 0
        rows.append({'Артикул_IH': art, 'Залишок': q, 'В_дорозі': in_transit})
    if not rows: return pd.DataFrame(columns=['Артикул_IH','Залишок','В_дорозі'])
    return pd.DataFrame(rows).groupby('Артикул_IH').agg(

        Залишок=('Залишок','sum'), В_дорозі=('В_дорозі','sum')).reset_index()

def extract_bc(x):
    if not isinstance(x, str): return None
    m = re.search(r'Штрихкод[:\s]*(\d{10,14})', x)
    return m.group(1) if m else None

@st.cache_data
def load_barcodes():
    # Спробуємо кілька варіантів шляху для сумісності з Streamlit Cloud
    candidates = [
        Path(__file__).parent / "data" / "barcodes.csv",
        Path("data/barcodes.csv"),
        Path("/mount/src") / Path(__file__).parent.name / "data" / "barcodes.csv",
    ]
    p = None
    for c in candidates:
        if c.exists():
            p = c
            break
    if p is None:
        st.error("❌ Файл data/barcodes.csv не знайдено. Переконайтесь що він є в репозиторії у папці data/")
        st.stop()
    df = pd.read_csv(p)
    df.columns = ['Артикул_IH','Назва_bc','Штрихкод']
    df['Штрихкод']   = df['Штрихкод'].astype(str).str.strip()
    df['Артикул_IH'] = df['Артикул_IH'].astype(str).str.strip()
    return dict(zip(df['Штрихкод'], df['Артикул_IH']))

# ── Uploaders ─────────────────────────────────────────────────────
st.markdown('<div class="sh"><div class="dot" style="background:#0ea5e9"></div><h3>Завантаження файлів</h3></div>', unsafe_allow_html=True)
c1,c2,c3,c4 = st.columns(4)
c5,c6,c7 = st.columns(3)
with c1: f_site   = st.file_uploader("🌐 Сайт (1С: продажі + залишки)", type=["xlsx","xls"], help="Листи: 'продажі дані' та 'наявність на складі'")
with c2: f_stores = st.file_uploader("🏪 Магазини (1С: продажі + залишки)", type=["xlsx","xls"], help="Листи: 'продажі магазини' та 'наявність на магазинах'")
with c3: f_rrp    = st.file_uploader("💲 РРЦ", type=["xlsx","xls","csv"])
with c4: f_ih     = st.file_uploader("📘 Прайс iHerb", type=["xlsx","xls","csv"])
with c5: f_vw     = st.file_uploader("📗 Прайс VitaWorld", type=["xls","xlsx","csv"])
with c6: f_dsn    = st.file_uploader("📙 Прайс DSN", type=["xlsx","xls","csv"])
with c7: f_atl    = st.file_uploader("📕 Прайс AtletikVit", type=["xlsx","xls","csv"])

required_missing = [n for n,f in [("Сайт",f_site),("Магазини",f_stores)] if not f]
if required_missing:
    st.info(f"Завантажте обов'язкові файли продажів. Ще не завантажено: **{', '.join(required_missing)}**")
    st.stop()

optional_missing = [n for n,f in [("РРЦ",f_rrp),("iHerb",f_ih),("VitaWorld",f_vw),
                                   ("DSN",f_dsn),("AtletikVit",f_atl)] if not f]
if optional_missing:
    st.warning(f"⚠️ Звіт сформується без: **{', '.join(optional_missing)}**. "
               f"Маржа/порівняння для цих джерел будуть недоступні, поки файли не завантажені.")

# ── Parse ─────────────────────────────────────────────────────────
with st.spinner("Обробка файлів..."):
    bc_map = load_barcodes()

    # Сайт (читання з кешем — повторний рендер не перечитує файл з диска)
    site_bytes = f_site.getvalue()
    df_site_s = cached_read_excel_sheet(site_bytes, 'продажі дані', None)
    df_site_z = cached_read_excel_sheet(site_bytes, 'наявність на складі', None)
    sales_site  = parse_sales(df_site_s, step=5, data_row=13, min_rent_pct=min_rent, n_months=recent_months)
    stock_site  = parse_stock(df_site_z, data_row=3)

    # Коригуємо середньоденні продажі сайту на фактичну к-сть днів наявності
    # (замість поділу на календарні 30 днів/місяць) — лист 'Залишки'
    excluded_anomaly = pd.DataFrame(columns=['Артикул_IH','Назва','Продано_всього'])
    if 'Залишки' in cached_excel_sheet_names(site_bytes):
        df_site_days = cached_read_excel_sheet(site_bytes, 'Залишки', None)
        avail_days_site = parse_availability_days(df_site_days, n_months=recent_months)
        sales_site = sales_site.merge(avail_days_site, on='Артикул_IH', how='left')
        # Якщо немає даних про дні наявності взагалі (артикул не зустрічався в Залишках) —
        # рахуємо по-старому (календарних 30 днів/місяць), бо це не аномалія, а просто відсутність рядка
        no_data_mask = sales_site['Днів_наявності'].isna()
        sales_site.loc[no_data_mask, 'Днів_наявності'] = sales_site.loc[no_data_mask, 'Місяців'] * 30

        # АНОМАЛІЯ: дні наявності = 0, але продажі > 0 (суперечливі дані 1С) — виключаємо з розрахунку,
        # показуємо окремо для ручної перевірки замість підстановки нереалістичного середньоденного
        anomaly_mask = (sales_site['Днів_наявності'] <= 0) & (sales_site['Продано_всього'] > 0)
        excluded_anomaly = sales_site.loc[anomaly_mask, ['Артикул_IH','Назва','Продано_всього']].copy()
        sales_site = sales_site[~anomaly_mask].copy()

        sales_site['Середньодень'] = sales_site['Продано_всього'] / sales_site['Днів_наявності']
        sales_site = sales_site.drop(columns=['Днів_наявності'])

    # Магазини (теж з кешем)
    store_bytes = f_stores.getvalue()
    df_store_s = cached_read_excel_sheet(store_bytes, 'продажі магазини', None)
    df_store_z = cached_read_excel_sheet(store_bytes, 'наявність на магазинах', None)
    sales_store = parse_sales(df_store_s, step=7, data_row=13, min_rent_pct=min_rent, n_months=recent_months)
    stock_store = parse_stock(df_store_z, data_row=2)
    # Для магазинів немає даних про фактичні дні наявності — лишаємо календарний розрахунок (n_m*30)

    # Зводимо
    df_sales = pd.concat([sales_site, sales_store], ignore_index=True).groupby('Артикул_IH').agg(
        Назва=('Назва','first'), Продано_всього=('Продано_всього','sum'),
        Місяців=('Місяців','max'), Середньодень=('Середньодень','sum')).reset_index()
    df_stock = pd.concat([stock_site, stock_store], ignore_index=True
                         ).groupby('Артикул_IH').agg(
        Залишок=('Залишок','sum'), В_дорозі=('В_дорозі','sum')).reset_index()

    # РРЦ
    # РРЦ — новий формат файлу: sku / price_before_coefficient (USD), без дублікатів і дат
    if f_rrp is not None:
        raw_rrp = cached_read_f(f_rrp.getvalue(), f_rrp.name); raw_rrp.columns = raw_rrp.columns.str.strip()
        ar = find_col(raw_rrp, ['sku','артикул'])
        pr = find_col(raw_rrp, ['price_before_coefficient'])
        if not pr:  # запасний варіант якщо колонку перейменують/файл іншого типу
            dr = find_col(raw_rrp, ['дата','period','період','date'])
            pr = find_col(raw_rrp, ['ціна','цена','price','special_price'])
            df_rrp = raw_rrp[[ar,dr,pr]].copy() if dr else raw_rrp[[ar,pr]].copy()
            df_rrp.columns = ['Артикул_IH','Дата_РРЦ','РРЦ_USD'] if dr else ['Артикул_IH','РРЦ_USD']
            df_rrp['Артикул_IH'] = df_rrp['Артикул_IH'].astype(str).str.strip()
            df_rrp['РРЦ_USD'] = pd.to_numeric(df_rrp['РРЦ_USD'], errors='coerce')
            if 'Дата_РРЦ' in df_rrp.columns:
                df_rrp['Дата_РРЦ'] = pd.to_datetime(df_rrp['Дата_РРЦ'], errors='coerce')
                df_rrp = df_rrp.sort_values('Дата_РРЦ').drop_duplicates('Артикул_IH', keep='last')
            else:
                df_rrp = df_rrp.drop_duplicates('Артикул_IH', keep='last')
            df_rrp = df_rrp[['Артикул_IH','РРЦ_USD']]
        else:
            df_rrp = raw_rrp[[ar, pr]].copy()
            df_rrp.columns = ['Артикул_IH','РРЦ_USD']
            df_rrp['Артикул_IH'] = df_rrp['Артикул_IH'].astype(str).str.strip()
            df_rrp['РРЦ_USD'] = pd.to_numeric(df_rrp['РРЦ_USD'], errors='coerce')
            df_rrp = df_rrp.drop_duplicates('Артикул_IH', keep='last')
    else:
        df_rrp = pd.DataFrame(columns=['Артикул_IH','РРЦ_USD'])

    # iHerb прайс
    if f_ih is not None:
        raw_ih = cached_read_f(f_ih.getvalue(), f_ih.name); raw_ih.columns = raw_ih.columns.str.strip()
        a_ih = find_col(raw_ih,['артикул','sku','article']); p_ih = find_col(raw_ih,['ціна','цена','price']); v_ih = find_col(raw_ih,['наявн','налич','availab'])
        df_ih_p = raw_ih[[a_ih,p_ih]].copy(); df_ih_p.columns = ['Артикул_IH','Ціна_IH_USD']
        df_ih_p['Наявність_IH'] = raw_ih[v_ih].values if v_ih else 'Є в наявності'
        df_ih_p['Артикул_IH'] = df_ih_p['Артикул_IH'].astype(str).str.strip()
        df_ih_p['Ціна_IH_USD'] = pd.to_numeric(df_ih_p['Ціна_IH_USD'], errors='coerce')
        df_ih_p = df_ih_p.drop_duplicates('Артикул_IH')
        nm_ih = find_col(raw_ih,['назв','name','номенклатура'])
        if nm_ih:
            # Правильний merge по артикулу - без зсуву після drop_duplicates
            names_df = raw_ih[[a_ih, nm_ih]].copy()
            names_df.columns = ['Артикул_IH','Назва_IH']
            names_df['Артикул_IH'] = names_df['Артикул_IH'].astype(str).str.strip()
            names_df = names_df.drop_duplicates('Артикул_IH', keep='first')
            df_ih_p = df_ih_p.merge(names_df, on='Артикул_IH', how='left')
    else:
        df_ih_p = pd.DataFrame(columns=['Артикул_IH','Ціна_IH_USD','Наявність_IH','Назва_IH'])

    # VitaWorld
    if f_vw is not None:
        try:
            raw_vw = cached_read_excel_sheet(f_vw.getvalue(), 0, None, engine='xlrd' if f_vw.name.endswith('.xls') else 'openpyxl')
            vw_rows = []
            for _, row in raw_vw.iterrows():
                art = row[2]
                if pd.notna(art) and isinstance(art,str) and art.strip() and art.strip()!='Артикул':
                    bc = row[10] if len(row)>10 else None
                    bc_s = ''
                    if pd.notna(bc):
                        bc_raw = str(bc).strip()
                        # Беремо лише числові штрихкоди EAN/UPC (не ASIN типу X000QWJYKZ)
                        digits_only = bc_raw.replace('.','').replace(',','')
                        if digits_only.isdigit() and 10 <= len(digits_only) <= 14:
                            try: bc_s = str(int(float(bc_raw)))
                            except: bc_s = ''
                        else:
                            bc_s = ''  # пропускаємо нечислові коди''
                    vw_rows.append({'Артикул_VW':art.strip(), 'Назва_VW': str(row[3]).strip() if pd.notna(row[3]) else '',
                                    'Ціна_VW_USD':pd.to_numeric(row[5],errors='coerce'), 'Штрихкод':bc_s})
            df_vw_p = pd.DataFrame(vw_rows)
            df_vw_p['Артикул_IH'] = df_vw_p['Штрихкод'].apply(lambda x: bc_map.get(str(x).strip()) if x else None)
        except Exception as e:
            st.warning(f"VitaWorld помилка: {e}"); df_vw_p = pd.DataFrame(columns=['Артикул_VW','Назва_VW','Ціна_VW_USD','Штрихкод','Артикул_IH'])
    else:
        df_vw_p = pd.DataFrame(columns=['Артикул_VW','Назва_VW','Ціна_VW_USD','Штрихкод','Артикул_IH'])

    # DSN
    if f_dsn is not None:
        raw_dsn = cached_read_f(f_dsn.getvalue(), f_dsn.name); raw_dsn.columns = raw_dsn.columns.str.strip()
        a_dsn = find_col(raw_dsn,['артикул']); p_dsn = find_col(raw_dsn,['цена','ціна','price'])
        av_dsn = find_col(raw_dsn,['наличие','наявн']); desc_c = find_col(raw_dsn,['описание товара (ua)','опис товара'])
        df_dsn_p = raw_dsn[[a_dsn,p_dsn]].copy(); df_dsn_p.columns = ['Артикул_DSN','Ціна_DSN_UAH']
        if av_dsn: df_dsn_p['Наявність_DSN'] = raw_dsn[av_dsn].values
        df_dsn_p['Артикул_DSN'] = df_dsn_p['Артикул_DSN'].astype(str).str.strip()
        df_dsn_p['Ціна_DSN_UAH'] = pd.to_numeric(df_dsn_p['Ціна_DSN_UAH'], errors='coerce')
        if desc_c:
            df_dsn_p['Штрихкод_DSN'] = raw_dsn[desc_c].apply(extract_bc)
            df_dsn_p['Артикул_IH'] = df_dsn_p['Штрихкод_DSN'].apply(lambda x: bc_map.get(str(x).strip()) if pd.notna(x) else None)
        else: df_dsn_p['Артикул_IH'] = None
    else:
        df_dsn_p = pd.DataFrame(columns=['Артикул_DSN','Ціна_DSN_UAH','Наявність_DSN','Артикул_IH'])

    # AtletikVit (заголовки таблиці на 2-му рядку файлу — header=1)
    if f_atl is not None:
        if f_atl.name.endswith('.csv'):
            raw_atl = pd.read_csv(io.BytesIO(f_atl.getvalue()), header=1)
        else:
            eng = 'xlrd' if f_atl.name.endswith('.xls') else 'openpyxl'
            raw_atl = pd.read_excel(io.BytesIO(f_atl.getvalue()), header=1, engine=eng)
        raw_atl.columns = raw_atl.columns.astype(str).str.strip()
        a_atl  = find_col(raw_atl, ['артикул'])
        p_atl  = find_col(raw_atl, ['ціна зі знижкою','цена со скидкой','discount'])
        if not p_atl: p_atl = find_col(raw_atl, ['ціна','цена','price'])
        bc_atl = find_col(raw_atl, ['штрихкод','barcode','ean','upc'])
        qty_atl= find_col(raw_atl, ['кількість на складі','количество','stock'])
        nm_atl = find_col(raw_atl, ['назва','название','name'])
        if not a_atl or not p_atl:
            st.error(f"❌ AtletikVit: не вдалось знайти колонки. Знайдені колонки файлу: {list(raw_atl.columns)}")
            st.stop()
        df_atl_p = raw_atl[[a_atl, p_atl]].copy()
        df_atl_p.columns = ['Артикул_ATL','Ціна_ATL_USD']
        df_atl_p['Артикул_ATL']   = df_atl_p['Артикул_ATL'].astype(str).str.strip()
        df_atl_p['Ціна_ATL_USD']  = pd.to_numeric(df_atl_p['Ціна_ATL_USD'], errors='coerce')
        if nm_atl:  df_atl_p['Назва_ATL']     = raw_atl[nm_atl].values
        if qty_atl: df_atl_p['Кількість_ATL'] = pd.to_numeric(raw_atl[qty_atl], errors='coerce').fillna(0)
        else:       df_atl_p['Кількість_ATL'] = 0
        if bc_atl:
            df_atl_p['Штрихкод_ATL'] = raw_atl[bc_atl].apply(
                lambda x: str(int(float(x))) if pd.notna(x) else '')
            df_atl_p['Артикул_IH'] = df_atl_p['Штрихкод_ATL'].apply(
                lambda x: bc_map.get(str(x).strip()) if x else None)
        else:
            df_atl_p['Артикул_IH'] = None
    else:
        df_atl_p = pd.DataFrame(columns=['Артикул_ATL','Ціна_ATL_USD','Назва_ATL','Кількість_ATL','Артикул_IH'])

# ── Demand ────────────────────────────────────────────────────────
df = df_sales.merge(df_stock, on='Артикул_IH', how='left')
df['Залишок']  = df['Залишок'].fillna(0)
df['В_дорозі'] = df['В_дорозі'].fillna(0) if 'В_дорозі' in df.columns else 0
df['Потреба'] = ((df['Середньодень'] * total_days) - df['Залишок']).clip(lower=0).round().astype(int)

# ABC
total_s = df['Продано_всього'].sum()
ds = df.sort_values('Продано_всього', ascending=False).copy()
ds['cum'] = ds['Продано_всього'].cumsum() / total_s if total_s > 0 else 0
ds['ABC'] = ds['cum'].apply(lambda x: 'A' if x<=.20 else ('B' if x<=.50 else 'C'))
df = df.merge(ds[['Артикул_IH','ABC']], on='Артикул_IH', how='left')
df['ABC'] = df['ABC'].fillna('C')
df = df[df['ABC'].isin(abc_inc)]
df['Середньомісячні'] = (df['Середньодень'] * 30).round(1)

# Merge prices
df = df.merge(df_ih_p[['Артикул_IH','Ціна_IH_USD','Наявність_IH']+(['Назва_IH'] if 'Назва_IH' in df_ih_p.columns else [])], on='Артикул_IH', how='left')
vw_m = df_vw_p[df_vw_p['Артикул_IH'].notna()][['Артикул_IH','Артикул_VW','Ціна_VW_USD']].drop_duplicates('Артикул_IH')
df = df.merge(vw_m, on='Артикул_IH', how='left')
dsn_cols = ['Артикул_IH','Артикул_DSN','Ціна_DSN_UAH']+(['Наявність_DSN'] if 'Наявність_DSN' in df_dsn_p.columns else [])
dsn_m = df_dsn_p[df_dsn_p['Артикул_IH'].notna()][dsn_cols].drop_duplicates('Артикул_IH')
df = df.merge(dsn_m, on='Артикул_IH', how='left')
atl_cols = ['Артикул_IH','Артикул_ATL','Ціна_ATL_USD','Кількість_ATL']
atl_m = df_atl_p[df_atl_p['Артикул_IH'].notna()][atl_cols].drop_duplicates('Артикул_IH')
df = df.merge(atl_m, on='Артикул_IH', how='left')
df = df.merge(df_rrp, on='Артикул_IH', how='left')
if 'Назва_IH' in df.columns:
    df['Назва'] = df['Назва_IH'].where(df['Назва_IH'].notna() & (df['Назва_IH']!=''), df['Назва'])

# Currency
if disp_cur=='UAH':
    df['p_IH']=df['Ціна_IH_USD']*usd_rate; df['p_VW']=df['Ціна_VW_USD']*usd_rate
    df['p_DSN']=df['Ціна_DSN_UAH']; df['p_RRP']=df['РРЦ_USD']*usd_rate
    df['p_ATL']=df['Ціна_ATL_USD']*usd_rate
else:
    df['p_IH']=df['Ціна_IH_USD']; df['p_VW']=df['Ціна_VW_USD']
    df['p_DSN']=df['Ціна_DSN_UAH']/usd_rate; df['p_RRP']=df['РРЦ_USD']
    df['p_ATL']=df['Ціна_ATL_USD']

# Availability
if only_avail:
    df['av_IH']  = df['Наявність_IH'].apply(is_avail_ih)
    df['av_VW']  = df['Артикул_VW'].notna() & df['p_VW'].notna()
    df['av_DSN'] = df.get('Наявність_DSN', pd.Series(False,index=df.index)).apply(is_avail_dsn) if 'Наявність_DSN' in df.columns else df['p_DSN'].notna()
    df['av_ATL'] = df['Кількість_ATL'].fillna(0) > 0
else:
    df['av_IH']=df['p_IH'].notna(); df['av_VW']=df['p_VW'].notna(); df['av_DSN']=df['p_DSN'].notna()
    df['av_ATL']=df['p_ATL'].notna()

df['eff_IH']  = df['p_IH'].where(df['av_IH']  & df['p_IH'].notna())
df['eff_VW']  = df['p_VW'].where(df['av_VW']  & df['p_VW'].notna())
df['eff_DSN'] = df['p_DSN'].where(df['av_DSN'] & df['p_DSN'].notna())
df['eff_ATL'] = df['p_ATL'].where(df['av_ATL'] & df['p_ATL'].notna())

# Best supplier
def choose(row):
    opts = {k:v for k,v in {'iHerb':row['eff_IH'],'VitaWorld':row['eff_VW'],'DSN':row['eff_DSN'],'AtletikVit':row['eff_ATL']}.items() if pd.notna(v)}
    if not opts: return pd.Series({'Постачальник':'—','Ціна_закупівлі':np.nan,'Сума':np.nan,'Рівні_ціни':False})
    best=min(opts,key=opts.get); bval=opts[best]
    near=sum(1 for v in opts.values() if v<=bval*1.05)>1
    qty=int(row['Потреба'])
    suma = bval*qty if qty > 0 else 0
    return pd.Series({'Постачальник':best,'Ціна_закупівлі':bval,'Сума':suma,'Рівні_ціни':near})

df[['Постачальник','Ціна_закупівлі','Сума','Рівні_ціни']] = df.apply(choose, axis=1)

# Margin
def get_mg(row, sup):
    pm={'iHerb':row['eff_IH'],'VitaWorld':row['eff_VW'],'DSN':row['eff_DSN'],'AtletikVit':row['eff_ATL']}
    cost=pm.get(sup); rrp=row['p_RRP']
    if pd.isna(rrp) or pd.isna(cost) or rrp<=0: return np.nan
    return round((rrp-cost)/rrp*100,1)

df['Маржа_%']   = df.apply(lambda r: get_mg(r,r['Постачальник']),axis=1)
df['Маржа_IH']  = df.apply(lambda r: get_mg(r,'iHerb'),axis=1)
df['Маржа_VW']  = df.apply(lambda r: get_mg(r,'VitaWorld'),axis=1)
df['Маржа_DSN'] = df.apply(lambda r: get_mg(r,'DSN'),axis=1)
df['Маржа_ATL'] = df.apply(lambda r: get_mg(r,'AtletikVit'),axis=1)
df['Маржа_ОК']     = df['Маржа_%'].apply(lambda x: pd.isna(x) or x>=margin_min)
df['Підняти_РРЦ']  = df['Маржа_%'].apply(lambda x: pd.notna(x) and x<margin_min)
df['Економія'] = df.apply(lambda r: (max([v for v in [r['eff_IH'],r['eff_VW'],r['eff_DSN'],r['eff_ATL']] if pd.notna(v)]+[0])-r['Ціна_закупівлі'])*int(r['Потреба']) if pd.notna(r['Ціна_закупівлі']) and r['Потреба']>0 else 0, axis=1)

# ── KPIs ──────────────────────────────────────────────────────────
found = df[df['Постачальник']!='—']
to_order = found[found['Потреба'] > 0]  # тільки те, що реально потрібно замовити
total_sum = to_order['Сума'].sum(); savings = to_order['Економія'].sum()
low_mg = df['Підняти_РРЦ'].sum(); avg_mg = df[df['Маржа_%'].notna()]['Маржа_%'].mean()

st.markdown('<div class="sh"><div class="dot" style="background:#f59e0b"></div><h3>Результати аналізу</h3></div>', unsafe_allow_html=True)
p_label = f"останні {recent_months} міс." if recent_months else "весь період"
st.caption(f"Продажі: **{p_label}** · Ігноруємо рент. < **{min_rent}%** · Покриття: **{total_days} дн.** (горизонт {horizon} + lead {lead_time} + запас {safety})")

st.markdown(f"""<div class="kpi-row">
<div class="kpi sky"><div class="l">Сума замовлення</div><div class="v">{total_sum:,.0f}</div><div class="s">{cur}</div></div>
<div class="kpi grn"><div class="l">До замовлення</div><div class="v">{len(to_order)}</div><div class="s">з {len(found)} знайдено · {len(df)} всього</div></div>
<div class="kpi vio"><div class="l">Економія</div><div class="v">{savings:,.0f}</div><div class="s">{cur} vs найдорожчий</div></div>
<div class="kpi amb"><div class="l">Середня маржа</div><div class="v">{avg_mg:.1f}%</div><div class="s">поріг {margin_min}%</div></div>
<div class="kpi red"><div class="l">⚠️ Підняти РРЦ</div><div class="v">{int(low_mg)}</div><div class="s">маржа < {margin_min}%</div></div>
</div>""", unsafe_allow_html=True)

sup_html = '<div class="sup-row">'
for sup,cls,ico in [('iHerb','ih','🔵'),('VitaWorld','vw','🟢'),('DSN','dsn','🟡'),('AtletikVit','atl','🟣')]:
    sub = to_order[to_order['Постачальник']==sup]
    if len(sub): sup_html += f'<div class="sc {cls}"><div class="sn">{ico} {sup}</div><div class="sv">{sub["Сума"].sum():,.0f} {cur}</div><div class="sd">{len(sub)} позицій · {int(sub["Потреба"].sum())} одиниць</div></div>'
sup_html += '</div>'
st.markdown(sup_html, unsafe_allow_html=True)

# ── Tables ────────────────────────────────────────────────────────
def build_tbl(data):
    t=pd.DataFrame()
    t['Артикул']=data['Артикул_IH']; t['Назва']=data['Назва']; t['ABC']=data['ABC']
    t['Залишок']=data['Залишок'].astype(int)
    t['🚚 В дорозі']=data['В_дорозі'].astype(int) if 'В_дорозі' in data.columns else 0
    t['Потреба']=data['Потреба']
    t['Постачальник']=data['Постачальник']
    t[f'Ціна ({cur})']=data['Ціна_закупівлі'].round(2); t[f'РРЦ ({cur})']=data['p_RRP'].round(2)
    t['Маржа %']=data['Маржа_%']; t['⚠️ РРЦ']=data['Підняти_РРЦ']; t[f'Сума ({cur})']=data['Сума'].round(2)
    t[f'iHerb ({cur})']=data['eff_IH'].round(2); t[f'VW ({cur})']=data['eff_VW'].round(2); t[f'DSN ({cur})']=data['eff_DSN'].round(2); t[f'ATL ({cur})']=data['eff_ATL'].round(2)
    t['≈ Рівні']=data['Рівні_ціни']; t['Маржа iH %']=data['Маржа_IH']; t['Маржа VW %']=data['Маржа_VW']; t['Маржа DSN %']=data['Маржа_DSN']; t['Маржа ATL %']=data['Маржа_ATL']
    t['Продано всього']=data['Продано_всього'].astype(int)
    t['Серед./міс.']=data['Середньомісячні'].round(1) if 'Середньомісячні' in data.columns else 0
    return t

cfg={'Назва':st.column_config.TextColumn(width='large'),'ABC':st.column_config.TextColumn(width='small'),
     '🚚 В дорозі':st.column_config.NumberColumn(width='small'),
     '⚠️ РРЦ':st.column_config.CheckboxColumn(width='small'),'≈ Рівні':st.column_config.CheckboxColumn(width='small'),
     **{f'{k} ({cur})':st.column_config.NumberColumn(format='%.2f') for k in ['Ціна','РРЦ','Сума','iHerb','VW','DSN','ATL']},
     **{f'Маржа{k}':st.column_config.NumberColumn(format='%.1f %%') for k in [' %',' iH %',' VW %',' DSN %',' ATL %']},
     'Серед./міс.':st.column_config.NumberColumn(format='%.1f')}

def show_tab(subset, sup=None, key_prefix='all'):
    if sup: subset=subset[subset['Постачальник']==sup]
    if not len(subset): st.info("Немає позицій"); return
    only_need = st.checkbox("Тільки артикули з потребою (Потреба > 0)", value=True, key=f"need_{key_prefix}")
    if only_need:
        subset = subset[subset['Потреба'] > 0]
    if not len(subset): st.info("Немає позицій з потребою"); return
    tbl = build_tbl(subset)
    st.dataframe(tbl, use_container_width=True, hide_index=True, column_config=cfg)
    # Підсумковий рядок
    total_row = {c: '' for c in tbl.columns}
    total_row['Артикул'] = '**Підсумок**'
    total_row['Потреба'] = int(subset['Потреба'].sum())
    total_row[f'Сума ({cur})'] = round(subset['Сума'].sum(), 2)
    st.dataframe(pd.DataFrame([total_row]), use_container_width=True, hide_index=True,
                 column_config=cfg)
    st.markdown(f"**{subset['Сума'].sum():,.2f} {cur}** · {len(subset)} позицій · {int(subset['Потреба'].sum())} одиниць")

n_low=int(df['Підняти_РРЦ'].sum()); n_miss=int((df['Постачальник']=='—').sum()); n_anom=len(excluded_anomaly)
tabs=st.tabs(["📋 Всі","🔵 iHerb","🟢 VitaWorld","🟡 DSN","🟣 AtletikVit",
              "💹 Порівняння цін",f"⚠️ Підняти РРЦ ({n_low})",f"❌ Не знайдено ({n_miss})",
              f"🔍 Дані з аномалією ({n_anom})"])
with tabs[0]: show_tab(found, key_prefix='all')
with tabs[1]: show_tab(found,'iHerb', key_prefix='ih')
with tabs[2]: show_tab(found,'VitaWorld', key_prefix='vw')
with tabs[3]: show_tab(found,'DSN', key_prefix='dsn')
with tabs[4]: show_tab(found,'AtletikVit', key_prefix='atl')
with tabs[5]:
    # Порівняння цін всіх постачальників
    st.caption("Всі позиції де є хоча б 1 постачальник — порівняння цін, РРЦ і маржі")
    only_need_cmp = st.checkbox("Тільки артикули з потребою (Потреба > 0)", value=True, key="need_cmp")
    price_df = df[df[['eff_IH','eff_VW','eff_DSN','eff_ATL']].notna().sum(axis=1) >= 1].copy()
    if only_need_cmp:
        price_df = price_df[price_df['Потреба'] > 0]
    if len(price_df):
        cmp = pd.DataFrame()
        cmp['Артикул']        = price_df['Артикул_IH']
        cmp['Назва']          = price_df['Назва']
        cmp['ABC']            = price_df['ABC']
        cmp['Залишок']        = price_df['Залишок'].astype(int)
        cmp['🚚 В дорозі']   = price_df['В_дорозі'].astype(int) if 'В_дорозі' in price_df.columns else 0
        cmp['Потреба']        = price_df['Потреба']
        cmp[f'iHerb ({cur})']     = price_df['eff_IH'].round(2)
        cmp[f'VitaWorld ({cur})'] = price_df['eff_VW'].round(2)
        cmp[f'DSN ({cur})']       = price_df['eff_DSN'].round(2)
        cmp[f'AtletikVit ({cur})']= price_df['eff_ATL'].round(2)
        cmp[f'РРЦ ({cur})']       = price_df['p_RRP'].round(2)
        cmp['Маржа iH %']     = price_df['Маржа_IH']
        cmp['Маржа VW %']     = price_df['Маржа_VW']
        cmp['Маржа DSN %']    = price_df['Маржа_DSN']
        cmp['Маржа ATL %']    = price_df['Маржа_ATL']
        cmp['✅ Найкраща']    = price_df['Постачальник']
        def calc_diff_pct(r):
            vals = [v for v in [r['eff_IH'], r['eff_VW'], r['eff_DSN'], r['eff_ATL']] if pd.notna(v)]
            if len(vals) < 2:
                return 0.0
            mx, mn = max(vals), min(vals)
            if mn <= 0:
                return 0.0
            return round((mx - mn) / mn * 100, 1)
        cmp['Різниця min-max %'] = price_df.apply(calc_diff_pct, axis=1)
        cmp = cmp.sort_values('Різниця min-max %', ascending=False)
        cmp_cfg = {
            'Назва': st.column_config.TextColumn(width='large'),
            'ABC':   st.column_config.TextColumn(width='small'),
            **{f'{k} ({cur})': st.column_config.NumberColumn(format='%.2f')
               for k in ['iHerb','VitaWorld','DSN','AtletikVit','РРЦ']},
            **{f'Маржа {k} %': st.column_config.NumberColumn(format='%.1f %%')
               for k in ['iH','VW','DSN','ATL']},
            'Різниця min-max %': st.column_config.NumberColumn(format='%.1f %%'),
        }
        st.dataframe(cmp, use_container_width=True, hide_index=True, column_config=cmp_cfg)
        # Підсумковий рядок — сума за обраним постачальником (найкращим)
        best_sum = price_df.apply(
            lambda r: r['Ціна_закупівлі']*int(r['Потреба']) if pd.notna(r['Ціна_закупівлі']) and r['Потреба']>0 else 0,
            axis=1).sum()
        total_row = {c: '' for c in cmp.columns}
        total_row['Артикул'] = '**Підсумок**'
        total_row['Потреба'] = int(price_df['Потреба'].sum())
        st.dataframe(pd.DataFrame([total_row]), use_container_width=True, hide_index=True, column_config=cmp_cfg)
        st.markdown(f"**{len(cmp)} позицій** · Потреба разом: **{int(price_df['Потреба'].sum())} од.** · "
                    f"Сума за найкращою ціною: **{best_sum:,.2f} {cur}**")
    else:
        st.info("Немає позицій з цінами постачальників")
with tabs[6]:
    low=df[df['Підняти_РРЦ']]
    if len(low): show_tab(low, key_prefix='low'); st.warning(f"⚠️ {len(low)} позицій: маржа нижче {margin_min}%. Розгляньте підняття РРЦ або зміну постачальника.")
    else: st.success(f"✅ Всі позиції з маржею ≥ {margin_min}%")
with tabs[7]:
    miss=df[df['Постачальник']=='—'][['Артикул_IH','Назва','ABC','Продано_всього','Залишок']]
    miss.columns=['Артикул','Назва','ABC','Продано','Залишок']
    st.dataframe(miss, use_container_width=True, hide_index=True)
with tabs[8]:
    st.caption("Сайт: товари з продажами, але 0 днів наявності за звітом 1С 'Залишки' за обраний період "
               "— суперечливі дані, тому виключені з автоматичного розрахунку потреби. Перевірте вручну.")
    if len(excluded_anomaly):
        anom_show = excluded_anomaly.rename(columns={'Продано_всього':'Продано (од.)'})
        st.dataframe(anom_show, use_container_width=True, hide_index=True)
        st.warning(f"⚠️ {len(excluded_anomaly)} артикулів сайту виключено з розрахунку потреби через аномалію в даних.")
    else:
        st.success("✅ Аномалій не знайдено — усі дані узгоджені.")

# ── Export ────────────────────────────────────────────────────────
st.markdown('<div class="sh"><div class="dot" style="background:#22c55e"></div><h3>Експорт замовлень</h3></div>', unsafe_allow_html=True)

def add_autofilter(ws, n_cols, n_rows):
    """Додає фільтри (випадаючі стрілки) на заголовки колонок Excel-листа."""
    from openpyxl.utils import get_column_letter
    last_col = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A1:{last_col}{n_rows}"
    # Закріплюємо заголовок щоб фільтри завжди були видимі при скролі
    ws.freeze_panes = "A2"

def sup_sheet(writer, sup, data):
    sub=data[(data['Постачальник']==sup) & (data['Потреба']>0)]
    if not len(sub): return
    out=pd.DataFrame({'Артикул (iHerb)':sub['Артикул_IH'],'Назва':sub['Назва'],'ABC':sub['ABC']})
    if sup=='VitaWorld' and 'Артикул_VW' in sub.columns: out['Артикул VW']=sub['Артикул_VW']
    if sup=='DSN' and 'Артикул_DSN' in sub.columns: out['Артикул DSN']=sub['Артикул_DSN']
    if sup=='AtletikVit' and 'Артикул_ATL' in sub.columns: out['Артикул ATL']=sub['Артикул_ATL']
    out['Залишок']=sub['Залишок'].astype(int)
    out['🚚 В дорозі']=sub['В_дорозі'].astype(int) if 'В_дорозі' in sub.columns else 0
    out['Потреба (шт.)']=sub['Потреба']
    out[f'Ціна ({cur})']=sub['Ціна_закупівлі'].round(2); out[f'РРЦ ({cur})']=sub['p_RRP'].round(2)
    out['Маржа %']=sub['Маржа_%']; out['⚠️ Підняти РРЦ']=sub['Підняти_РРЦ']
    out[f'Сума ({cur})']=sub['Сума'].round(2)
    n_data_rows = len(out)
    # Підсумковий рядок: сума потреби та сума замовлення
    total_row = {c: '' for c in out.columns}
    total_row['Артикул (iHerb)'] = 'ПІДСУМОК'
    total_row['Потреба (шт.)'] = int(out['Потреба (шт.)'].sum())
    total_row[f'Сума ({cur})'] = round(out[f'Сума ({cur})'].sum(), 2)
    out = pd.concat([out, pd.DataFrame([total_row])], ignore_index=True)
    out.to_excel(writer, index=False, sheet_name=sup)
    # Червона підсвітка рядків де треба підняти РРЦ + виділення підсумкового рядка
    try:
        from openpyxl.styles import PatternFill, Font
        ws_out = writer.sheets[sup]
        red_fill = PatternFill('solid', start_color='FFCCCC', fgColor='FFCCCC')
        bold_fill = PatternFill('solid', start_color='E8E8E8', fgColor='E8E8E8')
        raise_rrp = sub['Підняти_РРЦ'].values
        for row_i, is_low in enumerate(raise_rrp, start=2):
            if is_low:
                for col_i in range(1, len(out.columns)+1):
                    ws_out.cell(row=row_i, column=col_i).fill = red_fill
        # Підсумковий рядок жирним і сірим фоном
        total_row_idx = n_data_rows + 2  # +2: 1 за заголовок, 1 за 1-індексацію Excel
        for col_i in range(1, len(out.columns)+1):
            c = ws_out.cell(row=total_row_idx, column=col_i)
            c.fill = bold_fill
            c.font = Font(bold=True)
        add_autofilter(ws_out, len(out.columns), n_data_rows+1)
    except Exception:
        pass

def build_price_comparison_df():
    """Будує датафрейм порівняння цін усіх постачальників + РРЦ для експорту."""
    price_df = df[df[['eff_IH','eff_VW','eff_DSN','eff_ATL']].notna().sum(axis=1) >= 1].copy()
    if not len(price_df):
        return pd.DataFrame()
    cmp = pd.DataFrame()
    cmp['Артикул']            = price_df['Артикул_IH']
    cmp['Назва']              = price_df['Назва']
    cmp['ABC']                = price_df['ABC']
    cmp['Залишок']            = price_df['Залишок'].astype(int)
    cmp['В дорозі']           = price_df['В_дорозі'].astype(int) if 'В_дорозі' in price_df.columns else 0
    cmp['Потреба']            = price_df['Потреба']
    cmp[f'iHerb ({cur})']     = price_df['eff_IH'].round(2)
    cmp[f'VitaWorld ({cur})'] = price_df['eff_VW'].round(2)
    cmp[f'DSN ({cur})']       = price_df['eff_DSN'].round(2)
    cmp[f'AtletikVit ({cur})']= price_df['eff_ATL'].round(2)
    cmp[f'РРЦ ({cur})']       = price_df['p_RRP'].round(2)
    cmp['Маржа iHerb %']      = price_df['Маржа_IH']
    cmp['Маржа VitaWorld %']  = price_df['Маржа_VW']
    cmp['Маржа DSN %']        = price_df['Маржа_DSN']
    cmp['Маржа AtletikVit %'] = price_df['Маржа_ATL']
    cmp['Найкраща ціна у']    = price_df['Постачальник']
    def calc_diff_pct(r):
        vals = [v for v in [r['eff_IH'], r['eff_VW'], r['eff_DSN'], r['eff_ATL']] if pd.notna(v)]
        if len(vals) < 2:
            return 0.0
        mx, mn = max(vals), min(vals)
        if mn <= 0:
            return 0.0
        return round((mx - mn) / mn * 100, 1)
    cmp['Різниця min-max %'] = price_df.apply(calc_diff_pct, axis=1)
    return cmp.sort_values('Різниця min-max %', ascending=False)

def make_full():
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine='openpyxl') as w:
        rows=[{'Постачальник':s,
               'Позицій':len(found[(found['Постачальник']==s) & (found['Потреба']>0)]),
               'Одиниць':int(found[(found['Постачальник']==s) & (found['Потреба']>0)]['Потреба'].sum()),
               f'Сума ({cur})':round(found[(found['Постачальник']==s) & (found['Потреба']>0)]['Сума'].sum(),2),
               'Серед. маржа %':round(found[(found['Постачальник']==s) & (found['Потреба']>0)]['Маржа_%'].mean(),1)
               if found[(found['Постачальник']==s) & (found['Потреба']>0)]['Маржа_%'].notna().any() else None}
              for s in ['iHerb','VitaWorld','DSN','AtletikVit'] if len(found[(found['Постачальник']==s) & (found['Потреба']>0)])]
        pd.DataFrame(rows).to_excel(w,index=False,sheet_name='Зведення')

        # Лист порівняння цін
        cmp = build_price_comparison_df()
        if len(cmp):
            cmp.to_excel(w, index=False, sheet_name='Порівняння цін')
            try:
                add_autofilter(w.sheets['Порівняння цін'], len(cmp.columns), len(cmp)+1)
            except Exception:
                pass

        for s in ['iHerb','VitaWorld','DSN','AtletikVit']: sup_sheet(w,s,found)
        low_df=df[df['Підняти_РРЦ']]
        if len(low_df):
            low_out = pd.DataFrame({'Артикул':low_df['Артикул_IH'],'Назва':low_df['Назва'],
                          'Постачальник':low_df['Постачальник'],
                          f'Ціна ({cur})':low_df['Ціна_закупівлі'].round(2),
                          f'РРЦ ({cur})':low_df['p_RRP'].round(2),'Маржа %':low_df['Маржа_%'],
                          f'Різниця ({cur})':((low_df['p_RRP']-low_df['Ціна_закупівлі'])*-1).round(2)
                          })
            low_out.to_excel(w,index=False,sheet_name='⚠️ Підняти РРЦ')
            try:
                add_autofilter(w.sheets['⚠️ Підняти РРЦ'], len(low_out.columns), len(low_out)+1)
            except Exception:
                pass
        miss2=df[df['Постачальник']=='—'][['Артикул_IH','Назва','ABC','Продано_всього','Залишок']]
        if len(miss2):
            miss2.to_excel(w,index=False,sheet_name='Не знайдено')
            try:
                add_autofilter(w.sheets['Не знайдено'], len(miss2.columns), len(miss2)+1)
            except Exception:
                pass
    return buf.getvalue()

def make_sup_xlsx(sup):
    sub=found[found['Постачальник']==sup]
    if not len(sub): return None
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine='openpyxl') as w: sup_sheet(w,sup,found)
    return buf.getvalue()

date_s=datetime.now().strftime('%Y%m%d')
e1,e2,e3,e4,e5=st.columns(5)
with e1:
    d=make_sup_xlsx('iHerb')
    if d: st.download_button("⬇️ iHerb",d,f"order_iherb_{date_s}.xlsx",use_container_width=True)
with e2:
    d=make_sup_xlsx('VitaWorld')
    if d: st.download_button("⬇️ VitaWorld",d,f"order_vw_{date_s}.xlsx",use_container_width=True)
with e3:
    d=make_sup_xlsx('DSN')
    if d: st.download_button("⬇️ DSN",d,f"order_dsn_{date_s}.xlsx",use_container_width=True)
with e4:
    d=make_sup_xlsx('AtletikVit')
    if d: st.download_button("⬇️ AtletikVit",d,f"order_atl_{date_s}.xlsx",use_container_width=True)
with e5:
    st.download_button("📥 Повне замовлення",make_full(),f"procurement_{date_s}.xlsx",
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        use_container_width=True,type='primary')

st.caption(f"Курс: 1 USD = {usd_rate} UAH · Покриття: {total_days} дн. · Ігнор. рент. < {min_rent}% · Поріг маржі: {margin_min}% · ABC: {', '.join(abc_inc)}")
